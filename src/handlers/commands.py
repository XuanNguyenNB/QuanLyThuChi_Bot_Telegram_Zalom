"""
Basic command handlers: /start, /help, /today, /month, /insights, /export, /delete, /link
"""

import csv
import io
import logging

from telegram import Update
from telegram.ext import ContextTypes

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from ..models import get_session
from ..models import TransactionType
from ..services import (
    get_or_create_user,
    get_today_summary,
    get_month_summary,
    get_all_transactions,
    get_last_transaction,
    delete_transaction,
    get_spending_insights,
    link_user_by_phone,
)
from ..utils import (
    format_currency, 
    format_currency_full, 
    format_date, 
    format_datetime,
    get_vietnam_now,
)
from ..charts import generate_pie_chart, generate_bar_chart

logger = logging.getLogger(__name__)


async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle /start command"""
    user = update.effective_user
    
    async with await get_session() as session:
        await get_or_create_user(
            session,
            user_id=user.id,
            username=user.username,
            full_name=user.full_name
        )
    
    welcome_message = (
        f"Chào {user.first_name}! 👋\n\n"
        "Tôi là bot ghi chép chi tiêu. Gõ nhanh để ghi:\n"
        "• `cafe 50` → 50,000₫ (không cần gõ k)\n"
        "• `grab 35k` → 35,000₫\n"
        "• `tiền nhà 2tr` → 2,000,000₫\n\n"
        "📋 *Lệnh hữu ích:*\n"
        "/today • /month • /insights • /help\n\n"
        "💬 *Hỏi đáp:* Gõ tự nhiên như:\n"
        "_\"Tháng này chi bao nhiêu?\"_"
    )
    
    await update.message.reply_text(welcome_message, parse_mode="Markdown")


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle /help command"""
    help_text = (
        "📖 *HƯỚNG DẪN SỬ DỤNG*\n"
        "━━━━━━━━━━━━━━━\n\n"
        "💰 *Ghi chi tiêu - Gõ tự nhiên:*\n"
        "```\n"
        "cafe 50      → 50,000₫\n"
        "grab 35k     → 35,000₫\n"
        "tiền nhà 2tr → 2,000,000₫\n"
        "```\n\n"
        "📈 *Ghi thu nhập:*\n"
        "```\n"
        "bán hàng 350      → Thu 350,000₫\n"
        "lương 15tr        → Thu 15,000,000₫\n"
        "bán x 500 trừ vốn 200 → Thu 300,000₫\n"
        "```\n\n"
        "💬 *Hỏi đáp thông minh:*\n"
        "```\n"
        "Tháng này chi bao nhiêu?\n"
        "Tôi chi nhiều nhất vào gì?\n"
        "```\n\n"
        "📋 *Các lệnh:*\n"
        "• /today → Chi tiêu hôm nay\n"
        "• /month → Chi tiêu tháng\n"
        "• /insights → Phân tích thông minh\n"
        "• /edit → Sửa giao dịch (chọn ngày → giao dịch)\n"
        "• /delete → Xóa giao dịch gần nhất\n"
        "• /ghilai → Ghi lại giao dịch cho ngày trước\n"
        "• /export → Xuất file CSV\n\n"
        "💡 *Mẹo:* Không cần gõ 'k', bot tự hiểu!\n"
        "`50 cafe` = `50k cafe` = 50,000₫"
    )
    await update.message.reply_text(help_text, parse_mode="Markdown")


async def today_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle /today command - show today's summary"""
    user = update.effective_user
    
    try:
        async with await get_session() as session:
            db_user = await get_or_create_user(session, user.id, user.username, user.full_name)
            summary = await get_today_summary(session, db_user.id)
        
        if summary.transaction_count == 0:
            await update.message.reply_text("📭 Hôm nay chưa có giao dịch nào.")
            return
        
        # Separate income and expense transactions
        income_txs = [tx for tx in summary.transactions if tx.category and tx.category.type.value == "INCOME"]
        expense_txs = [tx for tx in summary.transactions if not tx.category or tx.category.type.value != "INCOME"]
        
        # Build message
        lines = [f"📅 *Hôm nay* ({format_date(get_vietnam_now())})\n"]
        
        # Income section
        lines.append(f"💰 *Thu: {format_currency_full(summary.total_income)}*")
        if income_txs:
            for tx in income_txs[:5]:
                lines.append(f"  + {format_currency(tx.amount)} - {tx.note or 'N/A'}")
            if len(income_txs) > 5:
                lines.append(f"  _... và {len(income_txs) - 5} giao dịch khác_")
        
        # Expense section
        lines.append(f"💸 *Chi: {format_currency_full(summary.total_expense)}*")
        if expense_txs:
            for tx in expense_txs[:5]:
                lines.append(f"  - {format_currency(tx.amount)} - {tx.note or 'N/A'}")
            if len(expense_txs) > 5:
                lines.append(f"  _... và {len(expense_txs) - 5} giao dịch khác_")
        
        lines.append("")  # Empty line
        
        # Balance
        balance = summary.total_income - summary.total_expense
        if balance >= 0:
            lines.append(f"📈 *Thặng dư: +{format_currency_full(balance)}*")
        else:
            lines.append(f"📉 *Thâm hụt: {format_currency_full(balance)}*")
        
        await update.message.reply_text("\n".join(lines), parse_mode="Markdown")
        
    except Exception as e:
        logger.error(f"Error in today_command: {e}")
        await update.message.reply_text("❌ Có lỗi xảy ra. Vui lòng thử lại sau.")


async def month_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle /month command - show monthly summary"""
    user = update.effective_user
    
    try:
        async with await get_session() as session:
            db_user = await get_or_create_user(session, user.id, user.username, user.full_name)
            summary = await get_month_summary(session, db_user.id)
        
        if summary.transaction_count == 0:
            await update.message.reply_text("📭 Tháng này chưa có giao dịch nào.")
            return
        
        # Build message
        now = get_vietnam_now()
        lines = [f"📊 *Tháng {now.month}/{now.year}*\n"]
        
        # Income section
        lines.append(f"💰 *Thu: {format_currency_full(summary.total_income)}*")
        
        lines.append("")  # Empty line
        
        # Expense section
        lines.append(f"💸 *Chi: {format_currency_full(summary.total_expense)}*")
        if summary.category_breakdown:
            lines.append(f"🏷️ Top danh mục:")
            for i, cat in enumerate(summary.category_breakdown[:5], 1):
                percent = (cat.total / summary.total_expense * 100) if summary.total_expense > 0 else 0
                lines.append(f"  {i}. {cat.category_name}: {format_currency_full(cat.total)} ({percent:.0f}%)")
        
        lines.append("")  # Empty line
        
        # Balance
        balance = summary.total_income - summary.total_expense
        if balance >= 0:
            lines.append(f"📈 *Thặng dư: +{format_currency_full(balance)}*")
        else:
            lines.append(f"📉 *Thâm hụt: {format_currency_full(balance)}*")
        
        await update.message.reply_text("\n".join(lines), parse_mode="Markdown")
        
        # Send Pie Chart if there are expenses
        if summary.total_expense > 0 and summary.category_breakdown:
            chart_data = [(cat.category_name, cat.total) for cat in summary.category_breakdown]
            chart_buf = generate_pie_chart(chart_data, f"Chi tiêu tháng {now.month}/{now.year}")
            if chart_buf:
                 await update.message.reply_photo(photo=chart_buf)
        
    except Exception as e:
        logger.error(f"Error in month_command: {e}")
        await update.message.reply_text("❌ Có lỗi xảy ra. Vui lòng thử lại sau.")


async def insights_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle /insights command - show spending insights"""
    user = update.effective_user
    
    try:
        async with await get_session() as session:
            db_user = await get_or_create_user(session, user.id, user.username, user.full_name)
            insights = await get_spending_insights(session, db_user.id)
        
        # Trend emoji
        trend_emoji = "📈" if insights.trend == "up" else "📉" if insights.trend == "down" else "➡️"
        
        lines = [
            "💡 *PHÂN TÍCH CHI TIÊU*",
            "",
            f"📊 *Tháng này:* {format_currency_full(insights.total_this_month)}",
            f"📊 *Tháng trước:* {format_currency_full(insights.total_last_month)}",
            f"{trend_emoji} *Xu hướng:* {'Tăng' if insights.trend == 'up' else 'Giảm' if insights.trend == 'down' else 'Ổn định'}",
            f"📅 *Trung bình/ngày:* {format_currency_full(insights.daily_average)}",
            "",
        ]
        
        if insights.top_categories:
            lines.append("🏷️ *Top 5 danh mục chi:*")
            for i, cat in enumerate(insights.top_categories[:5], 1):
                percent = (cat.total / insights.total_this_month * 100) if insights.total_this_month > 0 else 0
                lines.append(f"  {i}. {cat.category_name}: {format_currency_full(cat.total)} ({percent:.0f}%)")
            lines.append("")
        
        if insights.biggest_expense:
            lines.append(f"💸 *Chi lớn nhất:* {format_currency_full(insights.biggest_expense.amount)}")
            lines.append(f"   📝 {insights.biggest_expense.note or 'Không có ghi chú'}")
            lines.append("")
        
        lines.append(f"💬 *Gợi ý:* {insights.suggestion}")
        
        await update.message.reply_text("\n".join(lines), parse_mode="Markdown")
        
        # Comparison Chart (This Month vs Last Month)
        chart_data = [
            ("Tháng trước", insights.total_last_month),
            ("Tháng này", insights.total_this_month)
        ]
        chart_buf = generate_bar_chart(chart_data, "So sánh chi tiêu", y_label="VNĐ")
        if chart_buf:
            await update.message.reply_photo(photo=chart_buf)
        
    except Exception as e:
        logger.error(f"Error in insights_command: {e}")
        await update.message.reply_text("❌ Có lỗi xảy ra. Vui lòng thử lại sau.")


async def export_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle /export command - export transactions to CSV"""
    user = update.effective_user
    
    try:
        async with await get_session() as session:
            db_user = await get_or_create_user(session, user.id, user.username, user.full_name)
            transactions = await get_all_transactions(session, db_user.id)
        
        if not transactions:
            await update.message.reply_text("📭 Chưa có giao dịch nào để xuất.")
            return
        
        # Create CSV in memory
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow(["Ngày", "Số tiền", "Danh mục", "Ghi chú", "Loại"])
        
        # Data rows
        for tx in transactions:
            cat_name = tx.category.name if tx.category else "Khác"
            tx_type = "Thu" if (tx.category and tx.category.type == TransactionType.INCOME) else "Chi"
            writer.writerow([
                format_datetime(tx.date),
                tx.amount,
                cat_name,
                tx.note or "",
                tx_type
            ])
        
        # Send file
        output.seek(0)
        file_bytes = io.BytesIO(output.getvalue().encode('utf-8-sig'))  # UTF-8 BOM for Excel
        file_bytes.name = f"chi_tieu_{get_vietnam_now().strftime('%Y%m%d')}.csv"
        
        await update.message.reply_document(
            document=file_bytes,
            caption=f"📄 Xuất {len(transactions)} giao dịch thành công!"
        )
        
    except Exception as e:
        logger.error(f"Error in export_command: {e}")
        await update.message.reply_text("❌ Có lỗi xảy ra khi xuất file. Vui lòng thử lại sau.")


async def export_excel_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle /excel command - export transactions to Excel"""
    user = update.effective_user
    
    try:
        async with await get_session() as session:
            db_user = await get_or_create_user(session, user.id, user.username, user.full_name)
            transactions = await get_all_transactions(session, db_user.id)
        
        if not transactions:
            await update.message.reply_text("📭 Chưa có giao dịch nào để xuất.")
            return
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Chi tiêu"
        
        # Header
        headers = ["Ngày", "Số tiền", "Danh mục", "Ghi chú", "Loại"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for row_idx, tx in enumerate(transactions, 2):
            cat_name = tx.category.name if tx.category else "Khác"
            tx_type = "Thu" if (tx.category and tx.category.type == TransactionType.INCOME) else "Chi"
            ws.append([
                format_datetime(tx.date),
                tx.amount,
                cat_name,
                tx.note or "",
                tx_type
            ])
            
        # Save to buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        output.name = f"chi_tieu_{get_vietnam_now().strftime('%Y%m%d')}.xlsx"
        
        await update.message.reply_document(
            document=output,
            caption=f"📄 File Excel chi tiêu ({len(transactions)} giao dịch)"
        )
        
    except Exception as e:
        logger.error(f"Error in export_excel_command: {e}")
        await update.message.reply_text("❌ Có lỗi xảy ra. Vui lòng thử lại.")


async def delete_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle /delete command - delete last transaction"""
    user = update.effective_user
    
    try:
        async with await get_session() as session:
            db_user = await get_or_create_user(session, user.id, user.username, user.full_name)
            
            # Get last transaction
            last_tx = await get_last_transaction(session, db_user.id)
            
            if last_tx is None:
                await update.message.reply_text("❌ Không có giao dịch nào để xóa.")
                return
            
            # Store info before deleting
            amount = last_tx.amount
            note = last_tx.note or "Không có ghi chú"
            cat_name = last_tx.category.name if last_tx.category else "Khác"
            
            # Delete the transaction
            await delete_transaction(session, last_tx.id, db_user.id)
            
            # Get updated today's summary
            summary = await get_today_summary(session, db_user.id)
        
        await update.message.reply_text(
            f"🗑️ *Đã xóa giao dịch:*\n"
            f"💰 {format_currency_full(amount)}\n"
            f"📝 {note}\n"
            f"🏷️ {cat_name}\n"
            f"───────────────\n"
            f"💸 Tổng chi hôm nay: *{format_currency_full(summary.total_expense)}*",
            parse_mode="Markdown"
        )
        
    except Exception as e:
        logger.error(f"Error in delete_command: {e}")
        await update.message.reply_text("❌ Có lỗi xảy ra. Vui lòng thử lại.")


async def link_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle /link command - link with Zalo account by phone"""
    user = update.effective_user
    
    # Get phone from command args
    if not context.args or len(context.args) < 1:
        await update.message.reply_text(
            "📱 *Liên kết với Zalo*\n\n"
            "Để đồng bộ dữ liệu giữa Telegram và Zalo:\n"
            "1. Gõ: `/link 0901234567` (SĐT của bạn)\n"
            "2. Trên Zalo bot, gõ: `/link 0901234567`\n\n"
            "Sau khi liên kết, dữ liệu chi tiêu sẽ được đồng bộ!",
            parse_mode="Markdown"
        )
        return
    
    phone = context.args[0].strip()
    
    # Validate phone
    if not phone.isdigit() or len(phone) < 9:
        await update.message.reply_text("❌ Số điện thoại không hợp lệ.")
        return
    
    try:
        async with await get_session() as session:
            linked_user = await link_user_by_phone(session, phone, telegram_id=user.id)
            
            if linked_user is None:
                # Cannot link - telegram_id or phone already linked to another user
                await update.message.reply_text(
                    f"❌ *Không thể liên kết*\n\n"
                    f"SĐT {phone} hoặc tài khoản Telegram của bạn đã được liên kết với tài khoản khác.\n\n"
                    f"Mỗi SĐT chỉ có thể liên kết với một tài khoản Telegram và một tài khoản Zalo.",
                    parse_mode="Markdown"
                )
                return
            
            if linked_user.zalo_id:
                await update.message.reply_text(
                    f"✅ *Đã liên kết với Zalo!*\n"
                    f"📱 SĐT: {phone}\n\n"
                    f"Dữ liệu chi tiêu sẽ được đồng bộ giữa Telegram và Zalo.",
                    parse_mode="Markdown"
                )
            else:
                await update.message.reply_text(
                    f"📱 *Đã lưu SĐT:* {phone}\n\n"
                    f"Để đồng bộ với Zalo, hãy gõ `/link {phone}` trên Zalo bot.",
                    parse_mode="Markdown"
                )
                
    except Exception as e:
        logger.error(f"Error in link_command: {e}")
        await update.message.reply_text("❌ Có lỗi xảy ra. Vui lòng thử lại.")
